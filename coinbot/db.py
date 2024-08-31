import os
import sys
from datetime import date, datetime
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd
from box import Box
from loguru import logger
from tqdm import tqdm

from coinbot.metadata import coin_values, colors, countries
from coinbot.utils import convert_to_thousands


class DataBase:
    def __init__(self, file_path: str, latest_csv_path: str):
        """Constructor that reads the Excel file and sets up sheets."""
        self.file_path = file_path
        self.latest_csv_path = latest_csv_path

        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.deutschland_sheet = self.wb["Deutschland"]
        self.eu_sheet = self.wb["EU"]
        self.sonder_sheet = self.wb["Sondermünzen"]

        self.eu_df = self.setup_eu_dataframe()
        self.ger_df = self.setup_ger_dataframe()
        self.sonder_df = self.setup_sonder_dataframe()
        self.df = pd.concat([self.eu_df, self.ger_df, self.sonder_df])
        self.df.update(
            self.df.drop(columns=["Name", "Link"]).map(
                lambda x: x.lower() if isinstance(x, str) else x
            )
        )
        self.df = self.df.fillna(pd.NA).reset_index()
        if "index" in self.df.columns:
            self.df.drop(columns=["index"], inplace=True)
        self.align()
        self.save_df()

    def save_df(self):
        fp = os.path.join(
            os.path.dirname(__file__), os.pardir, "data", "latest_collection.csv"
        )
        self.df.to_csv(fp)

    def align(self):
        """
        Align the XLSM from Dropbox with the latest CSV used for this bot. This is useful
        to keep track which coin was collected when.
        """
        self.df.insert(6, "Collected", pd.NA)
        self.df.insert(7, "Collector", pd.NA)
        self.df.insert(5, "Created", pd.NA)
        self.df.insert(10, "Staged", pd.NA)
        self.latest_df = pd.read_csv(self.latest_csv_path).fillna(pd.NA)
        self.latest_df["Staged"] = self.latest_df["Staged"].fillna(False)

        added_coins = False  # tracks whether DF has new coins
        # Compare last version of DB with the one loaded from server
        for i, r in tqdm(
            self.df.iterrows(),
            total=len(self.df),
            desc="Aligning data",
            disable=not sys.stdout.isatty(),
        ):
            tdf = self.latest_df
            tdf = tdf[
                (tdf.Country == r.Country)
                & (tdf.Year == r.Year)
                & (tdf["Coin Value"] == r["Coin Value"])
                & (tdf["Special"] == r.Special)
                & (
                    (tdf["Source"].isna() & pd.isna(r.Source))
                    | (tdf["Source"] == r.Source)
                )
                & ((tdf["Name"].isna() & pd.isna(r.Name)) | (tdf["Name"] == r.Name))
            ]
            if len(tdf) > 1:
                logger.error(f"Multiple occurrences found: {tdf}")
                continue
            elif len(tdf) == 0:
                logger.warning(
                    f"Seems that coin ({r.Country}, {r.Year}, {r['Coin Value']}) was freshly added"
                )
                self.df.at[i, "Created"] = str(date.today().strftime("%d.%m.%Y"))
                if r.Status == "collected":
                    logger.info(
                        f"Fresh coin ({r.Country}, {r.Year}, {r['Coin Value']}) was already collected"
                    )
                    self.df.at[i, "Collected"] = str(date.today().strftime("%d.%m.%Y"))
                continue
            else:
                self.df.at[i, "Created"] = tdf.iloc[0].Created

            # Check whether status has changed
            matched_old_row = tdf.iloc[0]
            if r.Status == matched_old_row.Status:
                # Status did not change so we can copy over the old update date
                self.df.at[i, "Collected"] = matched_old_row.Collected
                self.df.at[i, "Collector"] = matched_old_row.Collector
            elif r.Status == "collected" and matched_old_row.Status != "collected":
                logger.info(
                    f"Coin ({r.Country}, {r.Year}, {r['Coin Value']}, {r.Source}, {r.Name}) was now collected by {matched_old_row.Collector}"
                )
                self.df.at[i, "Collected"] = str(date.today().strftime("%d.%m.%Y"))
                added_coins = True

                self.df.at[i, "Collector"] = matched_old_row.Collector
            elif matched_old_row.Status == "unavailable":
                # Status changed from unavailable to missing or sth else than collected
                pass
            else:
                raise ValueError(
                    f"Status divergence for old: {matched_old_row} vs. new: {r}"
                )

            self.df.at[i, "Staged"] = matched_old_row.Staged

        # Reset staged values if new coins were added to DB
        if added_coins:
            self.df["Staged"] = self.df["Staged"].fillna(False)
            self.df.loc[
                self.df["Staged"] & (self.df["Status"] != "collected"), "Collector"
            ] = np.nan
            self.df["Staged"] = False

    def get_status_diff(self, start: datetime, end: datetime):

        start_df = self.get_db_for_date(date=start)
        end_df = self.get_db_for_date(date=end)

        report_lines = []

        report_lines.append("🤑🪙Collection Change Status🤑🪙\n")
        report_lines.append("Color code:\n🟢Increase🟢\n🟡Static🟡\n🔴Decrease🔴\n\n")

        # Total coins info
        data = Box()
        for key, tdf in zip(["start", "end"], [start_df, end_df]):
            ## Global stats
            key_data = Box()
            key_data.coins = len(tdf)
            key_data.collected = len(tdf[tdf["Status"] == "collected"])
            key_data.special = len(tdf[tdf["Special"]])
            key_data.special_collected = len(
                tdf[(tdf["Status"] == "collected") & (tdf["Special"])]
            )
            key_data.total_ratio = key_data.collected / key_data.coins
            key_data.special_ratio = key_data.special_collected / key_data.special

            # Year stats
            for year in sorted(tdf["Year"].unique()):
                year_df = tdf[tdf["Year"] == year]
                key_year = Box()
                key_year.total = len(year_df)
                key_year.collected = len(year_df[year_df["Status"] == "collected"])
                key_year.ratio = (
                    key_year.collected / key_year.total if key_year.total > 0 else 0
                )

                key_data[str(year)] = key_year

            # Country stats
            for country in tdf["Country"].unique():
                country_df = tdf[tdf["Country"] == country]
                key_country = Box()
                key_country.total = len(country_df)
                key_country.collected = len(
                    country_df[country_df["Status"] == "collected"]
                )
                key_country.ratio = (
                    key_country.collected / key_country.total
                    if key_country.total > 0
                    else 0
                )
                key_data[country] = key_country

            for value in [f"{x} cent" for x in [1, 2, 5, 10, 20, 50]] + [
                f"{x} euro" for x in [1, 2]
            ]:
                value_df = tdf[tdf["Coin Value"] == value]
                key_value = Box()
                key_value.total = len(value_df)
                key_value.collected = len(value_df[value_df["Status"] == "collected"])
                key_value.ratio = (
                    key_value.collected / key_value.total if key_value.total > 0 else 0
                )
                key_data[value] = key_value

            data[key] = key_data
        trd = data.end.total_ratio - data.start.total_ratio
        srd = data.end.special_ratio - data.start.special_ratio

        # Formatting the total and special coins information
        report_lines.append(
            f"Total coins: {self._emojid(trd)}{trd:.2%}{self._emojid(trd)}: {data.start.total_ratio:.1%}➡️{data.end.total_ratio:.1%} ({data.start.coins}/{data.start.collected} > {data.end.coins}/{data.end.collected})\n"
        )
        report_lines.append(
            f"Special coins: {self._emojid(srd)}{srd:.2%}{self._emojid(srd)}: {data.start.special_ratio:.1%}➡️{data.end.special_ratio:.1%} ({data.start.special}/{data.start.special_collected} > {data.end.special}/{data.end.special_collected})\n"
        )

        report_lines.append("Years:")
        for year in sorted(end_df["Year"].unique()):
            year = str(year)
            yrd = data.end[year].ratio - data.start[year].ratio

            report_lines.append(
                f"{year}: {self._emojid(yrd)}{yrd:.2%}{self._emojid(yrd)}\n\t({data.start[year].ratio:.1%}➡️{data.end[year].ratio:.1%}, {data.start[year].collected}/{data.end[year].total}➡️{data.end[year].collected}/{data.end[year].total})"
            )

        report_lines.append("\nCountries:")
        for country in end_df["Country"].unique():
            crd = data.end[country].ratio - data.start[country].ratio
            report_lines.append(
                f"{country.capitalize()}: {self._emojid(crd)}{crd:.2%}{self._emojid(crd)}\n\t({data.start[country].ratio:.1%}➡️{data.end[country].ratio:.1%}, {data.start[country].collected}/{data.start[country].total}➡️{data.end[country].collected}/{data.end[country].total})"
            )

        # Generating report by Coin value
        report_lines.append("\nCoins:")  # Add a newline for separation
        for value in [f"{x} cent" for x in [1, 2, 5, 10, 20, 50]] + [
            f"{x} euro" for x in [1, 2]
        ]:
            vrd = data.end[value].ratio - data.start[value].ratio
            report_lines.append(
                f"{value}: {self._emojid(vrd)}{vrd:.2%}{self._emojid(vrd)}\n\t({data.start[value].ratio:.1%}➡️{data.end[value].ratio:.1%}, {data.start[value].collected}/{data.start[value].total}➡️{data.end[value].collected}/{data.end[value].total})"
            )

        # Joining report lines into a single string
        report = "\n".join(report_lines)
        return report

    def get_db_for_date(self, date: Optional[datetime] = None):
        df = self.df[self.df["Status"] != "unavailable"].copy()
        df["CreatedDate"] = pd.to_datetime(df["Created"], errors="coerce")
        df["CollectedDate"] = pd.to_datetime(df["Collected"], errors="coerce")
        if date is None:
            return df

        # Remove coins that were added to DB after the date
        df = df[df["CreatedDate"] <= date]
        # Coins that were collected after that date are changed to missing
        df.loc[df["CollectedDate"] > date, "Status"] = "missing"
        return df

    def get_status(self, msg: str):
        """
        Prints the database collection status report.

        Args:
            msg: The user message, can be Case 1, Case 2 or Case 3:

                Case 1: `Status` just gives the current status of DB
                Case 2: `Status DATE` gives the DB status at a specific date
                Case 3: `Status Diff DATE DATE` gives the delta across two timepoints
        Returns:
            A report describing the status. Or a error msg
        """
        words = msg.strip().split(" ")
        if msg.startswith("status diff"):
            # Case 3
            if not len(words) == 4:
                return "Need request in format `Status Diff 01.01.2024 01.08.2024`"
            start = datetime.strptime(words[-2], "%d.%m.%Y")
            end = datetime.strptime(words[-1], "%d.%m.%Y")
            report = self.get_status_diff(start=start, end=end)
            return report
        elif len(words) != 2:
            # Case 1 (default)
            df = self.get_db_for_date()
        else:
            given_date = datetime.strptime(words[-1], "%d.%m.%Y")
            df = self.get_db_for_date(date=given_date)

        assert (
            len(df[(df.Status == "collected") & (df.Staged == True)]) == 0
        ), "Some coin is collected AND staged"

        report_lines = []
        date_str = "Today" if len(words) == 1 else given_date
        report_lines.append(
            f"**🤑🪙 Collection Status as of {date_str} 🤑🪙**\n(Results including staged coins in brackets)\n"
        )
        report_lines.append(
            "Color code: 100% -> ✅ >90% -> 🟢 >80% -> 🟣 >70% -> 🔵 >60% -> ⚪ >50% -> 🟡 >40% -> 🟠 >30% -> 🔴 >20% -> 🟤 >10% -> ⚫ >0% -> ✖️ 0% -> 0️"
        )

        # Total coins info
        total_coins = len(df)
        special = len(df[df["Special"]])
        if total_coins == 0:
            report_lines.append("No data for this date. Pick a newer date")
            return "\n".join(report_lines)

        collected = len(df[df["Status"] == "collected"])
        stag = len(df[df["Staged"] == True])

        speccol = len(df[(df["Status"] == "collected") & (df["Special"])])
        tr = collected / total_coins
        trs = (collected + stag) / total_coins
        sr = speccol / special

        # Formatting the total and special coins information
        report_lines.append(
            f"**{self._emoji(tr)}({self._emoji(trs)}) Total coins: {total_coins}, done: {collected}({collected+stag}) {tr:.2%} ({trs:.2%})**"
        )
        report_lines.append(
            f"**{self._emoji(sr)}Special coins: {special}, done: {speccol} ({sr:.2%})**\n"
        )

        # Generating report by Year
        report_lines.append("Year:")  # Add a newline for separation
        for year in sorted(df["Year"].unique()):
            year_df = df[df["Year"] == year]
            tot = len(year_df)
            col = len(year_df[year_df["Status"] == "collected"])
            stag = len(year_df[year_df.Staged == True])

            fra = col / tot if tot > 0 else 0
            fras = (col + stag) / tot if tot > 0 else 0
            report_lines.append(
                f"{self._emoji(fra)}({self._emoji(fras)}) {year}: {fra:.2%} ({fras:.2%}) - {col}({col+stag}) / {tot}"
            )

        # Generating report by Country
        report_lines.append("\nCountries:")
        for country in df["Country"].unique():
            country_df = df[df["Country"] == country]
            tot = len(country_df)
            col = len(country_df[country_df["Status"] == "collected"])
            stag = len(country_df[country_df.Staged == True])

            fra = col / tot if tot > 0 else 0
            fras = (col + stag) / tot if tot > 0 else 0
            report_lines.append(
                f"{self._emoji(fra)}({self._emoji(fras)}) {country.capitalize()}: {fra:.2%} ({fras:.2%}) - {col}({col+stag}) / {tot}"
            )

        # Generating report by Coin value
        report_lines.append("\nCoin Value:")  # Add a newline for separation
        for value in [f"{x} cent" for x in [1, 2, 5, 10, 20, 50]] + [
            f"{x} euro" for x in [1, 2]
        ]:
            value_df = df[df["Coin Value"] == value]
            tot = len(value_df)
            col = len(value_df[value_df["Status"] == "collected"])
            stag = len(value_df[value_df.Staged == True])
            fra = col / tot if tot > 0 else 0
            fras = (col + stag) / tot if tot > 0 else 0
            report_lines.append(
                f"{self._emoji(fra)}({self._emoji(fras)}) {value}: {fra:.2%} ({fras:.2%}) - {col}({col+stag}) / {tot}"
            )

        # Joining report lines into a single string
        report = "\n".join(report_lines)
        return report

    def status_delta(self, year: int, value: str, country: str):
        """
        Sends a collection status update message to the user based on the
        information of the just-collected coin.
        """

        report_lines = ["📈Updated Stats (including staged coins!)📈\n"]
        df = self.get_db_for_date()

        def add_change(df: pd.DataFrame, msg: str):
            total_coins = len(df)
            collected = len(df[df["Status"] == "collected"])
            staged = len(df[df["Staged"] == True])
            assert (
                len(df[(df.Status == "collected") & (df.Staged == True)]) == 0
            ), "Some coin is collected AND staged"
            tro, trn = ((collected + staged - 1) / total_coins), (
                (collected + staged) / total_coins
            )
            emo, emn = self._emoji(tro), self._emoji(trn)
            report_lines.append(f"{msg}: From {emo}{tro:.3%} ➡️ {emn}{trn:.3%}")

        # 1. Overall change
        add_change(df, msg="Total")
        # 2. Country change
        add_change(df[df.Country == country], msg=f"{country.capitalize()}")
        # 3. Year change
        add_change(df[df.Year == year], msg=f"{year}")
        # 4. Coin value change
        add_change(df[df["Coin Value"] == value], msg=f"{value}")
        report = "\n".join(report_lines)
        return report

    def cell_status(self, cell):
        """Determine the collection status based on the cell color."""
        # Assuming default colors for collected, uncollected, and unavailable
        fill_color = cell.fill.start_color.index
        if fill_color not in colors.keys():
            logger.warning(f"Unknown cell color {fill_color} in {cell}")
        return colors.get(fill_color, "unknown")

    def setup_eu_dataframe(self):
        """Setup a dataframe for the EU sheet."""
        rows = list(self.eu_sheet.iter_rows(min_row=1))
        data = []
        num_countries = 0
        for i, row in enumerate(rows):
            if row[1].value in countries:
                # This row contains country names
                country = row[1].value
                # The next row contains the years
                year_row = rows[i + 1]
                years = []
                for cell in year_row[1:]:
                    if isinstance(cell.value, int) and 1999 <= cell.value <= 2030:
                        years.append(cell.value)
                    elif cell.value is None:
                        break  # Stop if the year is None
                # Loop over the years and the next 8 rows for coin values
                for year_idx, year in enumerate(years):
                    for value_idx, coin_value in enumerate(coin_values):
                        coin_row = rows[
                            (num_countries * 11) + value_idx + 2
                        ]  # Offset by 2 to skip country and year rows
                        cell = coin_row[
                            year_idx + 1
                        ]  # Offset by 1 to skip the coin value column
                        amount = (
                            cell.value if cell.value not in [None, "---", "???"] else 0
                        )
                        status = self.cell_status(cell)
                        data.append([country, year, coin_value, amount, status])
                num_countries += 1

        # Create DataFrame from data
        df = pd.DataFrame(
            data, columns=["Country", "Year", "Coin Value", "Amount", "Status"]
        )
        df["Amount"] = df["Amount"].apply(convert_to_thousands).astype(int)
        df["Coin Value"] = df["Coin Value"].str.lower()
        df["Special"] = False
        return df

    def setup_ger_dataframe(self):
        """Setup a dataframe for the Deutschland (Germany) sheet."""
        data = []
        rows = list(self.deutschland_sheet.iter_rows(min_row=1))

        # Define the start row for each 5-year block
        year_blocks = {2002: 0, 2007: 12, 2012: 24, 2017: 36, 2022: 48}

        for start_year, year_row in year_blocks.items():
            source_row = year_row + 1
            coin_value_rows = range(source_row + 1, source_row + 9)

            for year_idx in range(5):  # For each year in the block
                year = start_year + year_idx
                # Extract prägestätte marks and associate with years
                source_cell = rows[source_row][year_idx * 5 + 1 : year_idx * 5 + 6]
                sources = [cell.value for cell in source_cell]

                for value_idx, coin_row in enumerate(coin_value_rows):
                    coin_value = coin_values[value_idx]
                    for source_idx, source in enumerate(sources):
                        cell = rows[coin_row][year_idx * 5 + 1 + source_idx]
                        amount = (
                            cell.value if cell.value not in [None, "---", "???"] else 0
                        )

                        status = self.cell_status(cell)
                        data.append(
                            ["Germany", year, source, coin_value, amount, status]
                        )

        # Create DataFrame from data
        df = pd.DataFrame(data)
        df.columns = ["Country", "Year", "Source", "Coin Value", "Amount", "Status"]
        df["Amount"] = df["Amount"].apply(convert_to_thousands).astype(int)
        df["Coin Value"] = df["Coin Value"].str.lower()
        df["Special"] = False
        return df

    def setup_sonder_dataframe(self):
        """Setup a dataframe for the Sondermünzen sheet."""
        rows = list(self.sonder_sheet.iter_rows(min_row=1))
        data = []

        # Extract data for unstructured sondermünzen. Everything in the sheet has been collected.
        for i, row in enumerate(rows):
            if i < 2:
                continue

            name = row[0].value
            country = row[1].value
            year = row[2].value
            amount = int(row[3].value * 1000)  # Convert to thousands
            source = row[5].value
            cs = row[5].value is not None
            link = row[7].value
            desc = row[8].value
            state = self.cell_status(row[0])

            data.append(
                [name, country, year, "2 euro", source, amount, state, cs, desc, link]
            )

        # Create DataFrame from data
        df = pd.DataFrame(
            data,
            columns=[
                "Name",
                "Country",
                "Year",
                "Coin Value",
                "Source",
                "Amount",
                "Status",
                "Country-specific",
                "Description",
                "Link",
            ],
        )
        df["Coin Value"] = df["Coin Value"].str.lower()
        df["Special"] = True
        return df

    def _emojid(self, delta: float):
        if delta > 0:
            return "🟢"
        elif delta < 0:
            return "🔴"
        else:
            return "🟡"

    def _emoji(self, fraction: float) -> str:
        """Returns an emoji based on the fraction collected, one per decile."""

        if fraction < 0 or fraction > 1:
            return "❔"
        elif fraction == 1:
            return "✅"
        elif fraction >= 0.9:
            return "🟢"
        elif fraction >= 0.8:
            return "🟣"
        elif fraction >= 0.7:
            return "🔵"
        elif fraction >= 0.6:
            return "⚪"
        elif fraction >= 0.5:
            return "🟡"
        elif fraction >= 0.4:
            return "🟠"
        elif fraction >= 0.3:
            return "🔴"
        elif fraction >= 0.2:
            return "🟤"
        elif fraction >= 0.1:
            return "⚫"
        elif fraction > 0.0:
            return "✖️"
        elif fraction == 0:
            return "0️"
        else:
            return "❔"
