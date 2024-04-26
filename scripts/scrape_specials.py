import os
import re
from collections import Counter
from copy import deepcopy
from itertools import cycle, islice

import pandas as pd
import typer
from loguru import logger
from openpyxl import load_workbook
from wikitable import get_wikitable_and_imgs

from coinbot.formatting import fix_string, get_years, non_alphabetic
from coinbot.metadata import country_ger2eng


def main(
    filepath: str = typer.Option(
        ..., "--output", "-o", help="Path to the output .csv file"
    )
):
    dfs = get_wikitable_and_imgs()
    all_data = []
    entry_buffer = []  # Buffer to store entries until a description is found
    desc = ""
    years = []
    for i in range(len(dfs)):
        if i < 2:
            continue

        # Hack to avoid weirdly formatted columns
        tdf = dfs[i]
        tdf.to_csv("tmp.csv", encoding="UTF-8")
        tdf = pd.read_csv("tmp.csv", encoding="UTF-8", index_col=0)
        os.remove("tmp.csv")
        tdf.columns = [re.sub("[^A-Za-z0-9 \-]+", "", col) for col in tdf.columns]

        if "Land" not in tdf.columns:
            continue
        year = 2004 + len(years)
        years.append(year)
        for j, row in tdf.iterrows():
            if j == len(tdf) - 1:
                # We hit the last and empty row of the DF. Empty buffer if necessary
                if entry_buffer:
                    # Apply the last description to all buffer entries before ending the loop
                    for entry in entry_buffer:
                        entry.update({"Details": desc})
                        all_data.append(entry)
                break

            if (
                pd.isna(row["Ausgabedatum"])
                and not row["Land"].startswith("Beschreibung")
                or (pd.isna(row["Anlass"]) and pd.isna(row["Auflage"]))
            ):
                continue  # Non-coin-row

            if "Beschreibung" in row["Land"]:
                raw_desc = row["Land"].split("Beschreibung")[-1]
                desc = (
                    fix_string(raw_desc)
                    if non_alphabetic(raw_desc)
                    else fix_string(raw_desc.split(" ", 1)[1])
                )
                # Apply description to all buffered entries
                for entry in entry_buffer:
                    entry.update({"Details": fix_string(desc)})
                    all_data.append(entry)
                # Clear the buffer after processing
                entry_buffer = []
                continue

            name = fix_string(row["Anlass"])
            country = fix_string(row["Land"]).strip()
            # Count coins by spaces but ignoring San Marino
            num_entries = country.count(" ") - country.count("n M") + 1

            if num_entries > 1:
                countries = [c.strip() for c in country.split(" ")]
                coins = [
                    int(x.replace(".", "")) for x in row["Auflage"].strip().split(" ")
                ]
            else:
                countries = [country]
                coins = [int(row["Auflage"].replace(".", ""))]
            c_years = get_years(row["Ausgabedatum"])

            if c_years[0] != year:
                logger.warning(
                    f"Coin {name} from {country} issued on {row['Ausgabedatum']} in year {year}"
                )
            if len(c_years) < num_entries:
                c_years = list(islice(cycle(c_years), num_entries))

            for k in range(num_entries):
                amount = round(float(coins[k]) / 1000000, 3)
                base_entry = {
                    "Name der Münze": name,
                    "Herkunftsland": country_ger2eng[countries[k]],
                    "Ausgabejahr": year,
                    "Menge in Mill.": amount,
                    "Prägestätte": "",
                    "Wert": "2 euro",
                    "Link": row["imageurl"],
                }
                if countries[k] != "Deutschland":
                    entry_buffer.append(base_entry)
                else:
                    for stätte in ["A", "D", "G", "F", "J"]:
                        entry = deepcopy(base_entry)
                        entry["Prägestätte"] = stätte
                        entry["Menge in Mill."] = round(
                            amount / 5, 3
                        )  # this is a proxy
                        entry_buffer.append(entry)

    df = pd.DataFrame(all_data).sort_values(
        by=["Ausgabejahr", "Name der Münze", "Herkunftsland"], ascending=True
    )

    # Track whether the coin is country-specific or not
    name_counts = Counter(df["Name der Münze"])
    unique_names = [n for n in df["Name der Münze"] if name_counts[n] == 1]
    df.insert(
        4,
        "Landspezifisch?",
        df["Name der Münze"].apply(lambda x: x not in unique_names),
    )

    if filepath.endswith("xlsm"):
        wb = load_workbook(filepath, read_only=False, keep_vba=True)
        ws = wb["Sondermünzen"]

        # Collect existing names from the column (assumes 'Name der Münze' is in column A)
        existing_names = set()
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
            if row[0].value:
                existing_names.add(row[0].value)
        # Append new data only if the name is not already present
        for _, data_row in df.iterrows():
            name = data_row["Name der Münze"]
            if name not in existing_names:
                ws.append(data_row.tolist())

        wb.save(filepath.replace(".xlsm", "_new.xlsm"))
    elif filepath.endswith("csv"):
        if os.path.exists(filepath):
            df.to_csv(filepath.replace(".csv", "_new.csv"))
        else:
            df.to_csv(filepath)


if __name__ == "__main__":
    typer.run(main)
